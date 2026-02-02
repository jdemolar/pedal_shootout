from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
sheet = wb.active
sheet.title = "JHS Pedals"

# Headers
headers = [
    "Manufacturer",
    "Model",
    "Type of Effect",
    "Currently Available",
    "Inputs",
    "Outputs",
    "Power - Plug Size",
    "Power - Center Negative/Positive",
    "Power - Voltage",
    "Power - Current (mA)",
    "Dimensions (L x W x H)"
]

# Style headers
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col_num, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col_num)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Column widths
sheet.column_dimensions['A'].width = 15  # Manufacturer
sheet.column_dimensions['B'].width = 25  # Model
sheet.column_dimensions['C'].width = 18  # Type of Effect
sheet.column_dimensions['D'].width = 12  # Currently Available
sheet.column_dimensions['E'].width = 25  # Inputs
sheet.column_dimensions['F'].width = 25  # Outputs
sheet.column_dimensions['G'].width = 15  # Plug Size
sheet.column_dimensions['H'].width = 15  # Center +/-
sheet.column_dimensions['I'].width = 12  # Voltage
sheet.column_dimensions['J'].width = 15  # Current (mA)
sheet.column_dimensions['K'].width = 22  # Dimensions

# JHS Pedals data (based on information gathered)
pedals = [
    # Current Production Pedals
    ["JHS Pedals", "Morning Glory Clean", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Morning Glory V4", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "424 Gain Stage", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Notaklön", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Notaklön Pink", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Notaklön Splatter", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Notadümblë", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Colour Box V2", "Preamp", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS), XLR", "2.1mm", "Center Negative", "9V DC", "193", "5.7\" x 3.75\" x 1.85\""],
    ["JHS Pedals", "Colour Box 10", "Preamp", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS), XLR", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Hard Drive", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "78", "2.6\" x 4.8\" x 1.6\""],
    ["JHS Pedals", "Kilt V2", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Kilt 10", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Flight Delay", "Delay", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "The Violet", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "PackRat", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "PackRat White", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Bonsai", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Muffuletta", "Fuzz", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Pulp N Peel V4", "Compression", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS), XLR", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "AT+", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "15", "2.2\" x 4.8\" x 1.6\""],
    ["JHS Pedals", "Angry Charlie V3", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "PG-14", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Emperor V2", "Modulation", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Unicorn V2", "Modulation", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Kodiak", "Modulation", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Charlie Brown V4", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Clover", "Preamp", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS), XLR", "2.1mm", "Center Negative", "9V DC", "", "2.6\" x 4.8\" x 1.6\""],
    ["JHS Pedals", "Haunting Mids", "Utility", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Moonshine V2", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Cheese Ball", "Fuzz", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Milkman", "Delay", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Active A/B/Y", "Utility", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS) x2", "2.1mm", "Center Negative", "9V DC", "47", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "Prestige", "Utility", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Whitey Tighty", "Compression", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Tidewater", "Modulation", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Overdrive Preamp", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "51", "4.7\" x 3.7\" x 1.66\""],
    ["JHS Pedals", "Crayon", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Artificial Blonde", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    
    # 3 Series Pedals
    ["JHS Pedals", "3 Series Oil Can Delay", "Delay", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Rotary Chorus", "Modulation", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Tape Delay", "Delay", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Octave Reverb", "Reverb", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Reverb", "Reverb", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Screamer", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Overdrive", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Fuzz", "Fuzz", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Flanger", "Modulation", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Distortion", "Gain", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    ["JHS Pedals", "3 Series Harmonic Tremolo", "Modulation", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "70", "4.42\" x 2.38\" x 1.22\""],
    
    # EHX Collaboration
    ["JHS Pedals", "EHX by JHS - Big Muff 2", "Fuzz", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    
    # Utility Pedals
    ["JHS Pedals", "Summing Amp", "Utility", "Yes", "Mono (1/4\" TS) x2", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "19", "3.6\" x 1.5\" x 1\""],
    ["JHS Pedals", "Switchback", "Utility", "Yes", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "31", ""],
    
    # Discontinued Pedals
    ["JHS Pedals", "Double Barrel V4", "Multi Effects", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Sweet Tea V3", "Multi Effects", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Notaklön Blackout", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "SuperBolt V2", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC or 18V DC", "100", ""],
    ["JHS Pedals", "Spring Tank", "Reverb", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Big Muffuletta", "Fuzz", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Lizard Queen", "Fuzz", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Germanium Boost", "Utility", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Voice Tech", "Utility", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "Bat Sim", "Utility", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "The Bulb", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "VCR (Space Commander)", "Multi Effects", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "80", ""],
    ["JHS Pedals", "Panther Cub", "Delay", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    
    # ROSS Collaborations (Discontinued)
    ["JHS Pedals", "ROSS Compressor", "Compression", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "ROSS Distortion", "Gain", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "ROSS Chorus", "Modulation", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "ROSS Phaser", "Modulation", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
    ["JHS Pedals", "ROSS Fuzz", "Fuzz", "No", "Mono (1/4\" TS)", "Mono (1/4\" TS)", "2.1mm", "Center Negative", "9V DC", "", ""],
]

# Add data rows
cell_font = Font(name="Arial", size=10)
for row_num, pedal_data in enumerate(pedals, 2):
    for col_num, value in enumerate(pedal_data, 1):
        cell = sheet.cell(row=row_num, column=col_num)
        cell.value = value
        cell.font = cell_font
        cell.border = border
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# Freeze the header row
sheet.freeze_panes = 'A2'

# Set row height for header
sheet.row_dimensions[1].height = 45

wb.save('/mnt/user-data/outputs/JHS_Pedals_Database.xlsx')
print("JHS Pedals database created successfully!")
